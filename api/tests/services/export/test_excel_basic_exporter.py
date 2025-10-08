"""Tests for Excel Basic exporter."""

import pytest
from openpyxl import load_workbook
from io import BytesIO
from app.services.export.excel_basic_exporter import ExcelBasicExporter
from app.schemas.export import ExportInclude


@pytest.fixture
def sample_results():
    """Sample analysis results for testing."""
    return {
        "task_id": "t_test123456",
        "summary": {
            "nps": {
                "score": 25.5,
                "promoters": 30,
                "promoters_percentage": 50.0,
                "passives": 10,
                "passives_percentage": 16.7,
                "detractors": 20,
                "detractors_percentage": 33.3
            },
            "churn_risk": {
                "average": 0.35,
                "high_risk_count": 15,
                "high_risk_percentage": 25.0
            },
            "pain_points": [
                {"category": "Slow support", "count": 25, "percentage": 41.7},
                {"category": "High price", "count": 15, "percentage": 25.0}
            ]
        },
        "metadata": {
            "total_comments": 60,
            "processing_time_seconds": 12.5,
            "model_used": "gpt-4o-mini",
            "batches_processed": 2
        },
        "rows": [
            {
                "index": 0,
                "original_text": "Great service!",
                "nota": 9,
                "nps_category": "promoter",
                "sentiment": "positive",
                "language": "en",
                "churn_risk": 0.1,
                "pain_points": [],
                "emotions": {"satisfaccion": 0.9, "alegria": 0.7}
            },
            {
                "index": 1,
                "original_text": "Too slow",
                "nota": 3,
                "nps_category": "detractor",
                "sentiment": "negative",
                "language": "en",
                "churn_risk": 0.8,
                "pain_points": ["slow support"],
                "emotions": {"frustracion": 0.8, "enojo": 0.6}
            }
        ]
    }


class TestExcelBasicExporter:
    """Test suite for ExcelBasicExporter class."""

    def test_initialization(self):
        """Test ExcelBasicExporter initializes correctly."""
        exporter = ExcelBasicExporter()
        assert exporter.formatter is not None

    def test_get_media_type(self):
        """Test get_media_type returns correct MIME type."""
        exporter = ExcelBasicExporter()
        media_type = exporter.get_media_type()
        assert "spreadsheet" in media_type

    def test_export_summary(self, sample_results):
        """Test exporting summary only."""
        exporter = ExcelBasicExporter()
        content, filename = exporter.export(
            sample_results,
            ExportInclude.SUMMARY,
            "t_test123456"
        )

        # Verify content is bytes
        assert isinstance(content, bytes)

        # Load workbook and verify structure
        wb = load_workbook(BytesIO(content))
        assert "Resumen" in wb.sheetnames

        # Verify summary sheet has data
        ws = wb["Resumen"]
        assert ws.max_row > 1  # More than just headers

        # Verify filename format
        assert filename.startswith("analysis_t_test123456_")
        assert filename.endswith(".xlsx")

    def test_export_detailed(self, sample_results):
        """Test exporting detailed data only."""
        exporter = ExcelBasicExporter()
        content, filename = exporter.export(
            sample_results,
            ExportInclude.DETAILED,
            "t_test123456"
        )

        # Verify content is bytes
        assert isinstance(content, bytes)

        # Load workbook and verify structure
        wb = load_workbook(BytesIO(content))
        assert "Análisis Detallado" in wb.sheetnames

        # Verify detailed sheet has data
        ws = wb["Análisis Detallado"]
        assert ws.max_row >= 3  # Headers + at least 2 rows of data

    def test_export_all(self, sample_results):
        """Test exporting all data (5 sheets)."""
        exporter = ExcelBasicExporter()
        content, filename = exporter.export(
            sample_results,
            ExportInclude.ALL,
            "t_test123456"
        )

        # Verify content is bytes
        assert isinstance(content, bytes)

        # Load workbook and verify all sheets exist
        wb = load_workbook(BytesIO(content))
        assert "Resumen" in wb.sheetnames
        assert "Análisis Detallado" in wb.sheetnames
        assert "Emociones" in wb.sheetnames
        assert "Puntos de Dolor" in wb.sheetnames
        assert "Metadatos" in wb.sheetnames

        # Verify each sheet has data
        assert wb["Resumen"].max_row > 1
        assert wb["Análisis Detallado"].max_row >= 3
        assert wb["Emociones"].max_row >= 3
        assert wb["Puntos de Dolor"].max_row >= 3
        assert wb["Metadatos"].max_row >= 2

    def test_export_validates_results(self):
        """Test that export validates results structure."""
        exporter = ExcelBasicExporter()

        with pytest.raises(ValueError, match="Results missing required field"):
            exporter.export(
                {"invalid": "data"},
                ExportInclude.ALL,
                "t_test123456"
            )

    def test_export_empty_rows(self, sample_results):
        """Test exporting when no rows are present."""
        sample_results["rows"] = []
        exporter = ExcelBasicExporter()

        content, filename = exporter.export(
            sample_results,
            ExportInclude.ALL,
            "t_test123456"
        )

        # Should still create workbook with summary
        wb = load_workbook(BytesIO(content))
        assert "Resumen" in wb.sheetnames

    def test_filename_generation(self, sample_results):
        """Test filename includes task_id and timestamp."""
        exporter = ExcelBasicExporter()
        _, filename1 = exporter.export(sample_results, ExportInclude.SUMMARY, "t_abc")
        _, filename2 = exporter.export(sample_results, ExportInclude.SUMMARY, "t_xyz")

        # Different task IDs should generate different filenames
        assert "t_abc" in filename1
        assert "t_xyz" in filename2
        assert filename1 != filename2

    def test_pain_points_sheet_structure(self, sample_results):
        """Test pain points sheet has correct structure."""
        exporter = ExcelBasicExporter()
        content, _ = exporter.export(
            sample_results,
            ExportInclude.ALL,
            "t_test123456"
        )

        wb = load_workbook(BytesIO(content))
        ws = wb["Puntos de Dolor"]

        # Check headers exist
        assert ws.cell(1, 1).value is not None

        # Check data rows
        assert ws.max_row >= 3  # Header + 2 pain points

    def test_emotions_sheet_structure(self, sample_results):
        """Test emotions sheet has correct structure."""
        exporter = ExcelBasicExporter()
        content, _ = exporter.export(
            sample_results,
            ExportInclude.ALL,
            "t_test123456"
        )

        wb = load_workbook(BytesIO(content))
        ws = wb["Emociones"]

        # Check headers exist (Index + emotion columns)
        assert ws.cell(1, 1).value is not None

        # Check data rows (2 rows of data)
        assert ws.max_row >= 3  # Header + 2 data rows
