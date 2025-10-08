"""Summary sheet builder with styling."""

from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

from app.services.export.formatters.excel_styles import (
    ExcelColors, apply_header_style, apply_body_style, auto_adjust_column_width
)


class SummarySheetBuilder:
    """Builds styled summary sheet for Excel export."""

    def build(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Build summary sheet with key metrics and NPS chart.

        Args:
            wb: Workbook object
            results: Analysis results dictionary
        """
        ws = wb.create_sheet("Resumen")

        summary = results.get("summary", {})
        nps_data = summary.get("nps", {})
        churn_data = summary.get("churn_risk", {})
        metadata = results.get("metadata", {})

        # Add title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "Resumen de Análisis"
        apply_header_style(title_cell, ExcelColors.PRIMARY)

        # Add metrics
        metrics = [
            ("NPS Score", round(nps_data.get("score", 0), 1)),
            ("Total Comentarios", metadata.get("total_comments", 0)),
            ("Promotores", f"{nps_data.get('promoters', 0)} ({round(nps_data.get('promoters_percentage', 0), 1)}%)"),
            ("Pasivos", f"{nps_data.get('passives', 0)} ({round(nps_data.get('passives_percentage', 0), 1)}%)"),
            ("Detractores", f"{nps_data.get('detractors', 0)} ({round(nps_data.get('detractors_percentage', 0), 1)}%)"),
            ("Riesgo Promedio (%)", round(churn_data.get("average", 0) * 100, 1)),
            ("Alto Riesgo", f"{churn_data.get('high_risk_count', 0)} ({round(churn_data.get('high_risk_percentage', 0), 1)}%)"),
            ("Tiempo Procesamiento (s)", round(metadata.get("processing_time_seconds", 0), 1))
        ]

        # Write metrics with styling
        for row_idx, (metric, value) in enumerate(metrics, 3):
            metric_cell = ws.cell(row=row_idx, column=1, value=metric)
            value_cell = ws.cell(row=row_idx, column=2, value=value)

            apply_body_style(metric_cell, bold=True)
            apply_body_style(value_cell)

        # Add NPS distribution chart
        self._add_nps_chart(ws, nps_data, start_row=len(metrics) + 5)

        # Auto-adjust column widths
        auto_adjust_column_width(ws)

    def _add_nps_chart(self, ws, nps_data: Dict[str, Any], start_row: int) -> None:
        """
        Add NPS distribution pie chart.

        Args:
            ws: Worksheet object
            nps_data: NPS metrics dictionary
            start_row: Row to start chart data
        """
        # Add chart data
        ws.cell(row=start_row, column=1, value="Categoría")
        ws.cell(row=start_row, column=2, value="Cantidad")

        categories = [
            ("Promotores", nps_data.get("promoters", 0)),
            ("Pasivos", nps_data.get("passives", 0)),
            ("Detractores", nps_data.get("detractors", 0))
        ]

        for idx, (category, count) in enumerate(categories, 1):
            ws.cell(row=start_row + idx, column=1, value=category)
            ws.cell(row=start_row + idx, column=2, value=count)

        # Create pie chart
        chart = PieChart()
        chart.title = "Distribución NPS"
        chart.style = 10

        labels = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + 3)
        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # Position chart
        ws.add_chart(chart, f"D{start_row}")
