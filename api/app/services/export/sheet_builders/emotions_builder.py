"""Emotions sheet builder with styling."""

from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from app.services.export.formatters.excel_styles import (
    ExcelColors, apply_header_style, apply_body_style, auto_adjust_column_width
)
from app.services.export.formatters.dataframe_formatter import DataFrameFormatter


class EmotionsSheetBuilder:
    """Builds styled emotions sheet for Excel export."""

    def __init__(self):
        """Initialize emotions builder."""
        self.formatter = DataFrameFormatter()

    def build(self, wb: Workbook, rows_data: List[Dict[str, Any]]) -> None:
        """
        Build emotions analysis sheet with bar chart.

        Args:
            wb: Workbook object
            rows_data: List of row dictionaries
        """
        ws = wb.create_sheet("Emociones")

        # Create DataFrame
        df = self.formatter.create_emotions_dataframe(rows_data)

        # Write headers with styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            apply_header_style(cell, ExcelColors.SUCCESS)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_body_style(cell)

        # Add average emotions chart if data exists
        if len(df) > 0 and len(df.columns) > 1:
            self._add_emotions_chart(ws, df, start_row=len(df) + 4)

        # Auto-adjust column widths
        auto_adjust_column_width(ws)

    def _add_emotions_chart(self, ws, df, start_row: int) -> None:
        """
        Add bar chart showing average emotion scores.

        Args:
            ws: Worksheet object
            df: DataFrame with emotions
            start_row: Row to start chart
        """
        # Calculate averages for each emotion (excluding Index column)
        emotion_cols = [col for col in df.columns if col != "Index"]

        # Add chart data
        ws.cell(row=start_row, column=1, value="Emoción")
        ws.cell(row=start_row, column=2, value="Promedio (%)")

        for idx, emotion in enumerate(emotion_cols, 1):
            avg_value = df[emotion].mean()
            ws.cell(row=start_row + idx, column=1, value=emotion)
            ws.cell(row=start_row + idx, column=2, value=round(avg_value, 1))

        # Create bar chart
        chart = BarChart()
        chart.title = "Emociones Promedio"
        chart.style = 10
        chart.y_axis.title = "Porcentaje (%)"
        chart.x_axis.title = "Emoción"

        data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(emotion_cols))
        categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(emotion_cols))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, f"D{start_row}")
