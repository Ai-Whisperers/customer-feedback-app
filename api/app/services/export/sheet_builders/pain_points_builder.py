"""Pain points sheet builder with styling."""

from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

from app.services.export.formatters.excel_styles import (
    ExcelColors, apply_header_style, apply_body_style, auto_adjust_column_width
)
from app.services.export.formatters.dataframe_formatter import DataFrameFormatter


class PainPointsSheetBuilder:
    """Builds styled pain points sheet for Excel export."""

    def __init__(self):
        """Initialize pain points builder."""
        self.formatter = DataFrameFormatter()

    def build(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Build pain points sheet with horizontal bar chart.

        Args:
            wb: Workbook object
            results: Analysis results dictionary
        """
        ws = wb.create_sheet("Puntos de Dolor")

        # Create DataFrame
        df = self.formatter.create_pain_points_dataframe(results)

        # Write headers with styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            apply_header_style(cell, ExcelColors.WARNING)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_body_style(cell)

        # Add chart if data exists
        if len(df) > 1 or (len(df) == 1 and df.iloc[0]["Punto de Dolor"] != "Sin datos"):
            self._add_pain_points_chart(ws, df)

        # Auto-adjust column widths
        auto_adjust_column_width(ws)

    def _add_pain_points_chart(self, ws, df) -> None:
        """
        Add horizontal bar chart for pain points.

        Args:
            ws: Worksheet object
            df: DataFrame with pain points
        """
        # Take top 10 pain points for chart
        max_rows = min(10, len(df))

        # Create bar chart
        chart = BarChart()
        chart.type = "bar"  # Horizontal bars
        chart.title = "Top Puntos de Dolor"
        chart.style = 10
        chart.x_axis.title = "Frecuencia"
        chart.y_axis.title = "Punto de Dolor"

        # Reference data (column B: Frecuencia)
        data = Reference(ws, min_col=2, min_row=1, max_row=max_rows + 1)
        categories = Reference(ws, min_col=1, min_row=2, max_row=max_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Position chart to the right of data
        ws.add_chart(chart, "E2")
