"""Details sheet builder with styling."""

from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule

from app.services.export.formatters.excel_styles import (
    ExcelColors, apply_header_style, apply_body_style, auto_adjust_column_width
)
from app.services.export.formatters.dataframe_formatter import DataFrameFormatter


class DetailsSheetBuilder:
    """Builds styled details sheet for Excel export."""

    def __init__(self):
        """Initialize details builder."""
        self.formatter = DataFrameFormatter()

    def build(self, wb: Workbook, rows_data: List[Dict[str, Any]]) -> None:
        """
        Build detailed analysis sheet with conditional formatting.

        Args:
            wb: Workbook object
            rows_data: List of row dictionaries
        """
        ws = wb.create_sheet("Análisis Detallado")

        # Create DataFrame
        df = self.formatter.create_detailed_dataframe(rows_data)

        # Write headers with styling
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            apply_header_style(cell, ExcelColors.SECONDARY)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                apply_body_style(cell)

        # Apply conditional formatting to churn risk column
        churn_col_idx = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if "Riesgo" in col_name:
                churn_col_idx = col_idx
                break

        if churn_col_idx:
            self._apply_churn_risk_formatting(ws, churn_col_idx, len(df) + 1)

        # Auto-adjust column widths
        auto_adjust_column_width(ws)

    def _apply_churn_risk_formatting(self, ws, col_idx: int, max_row: int) -> None:
        """
        Apply color scale formatting to churn risk column.

        Args:
            ws: Worksheet object
            col_idx: Column index for churn risk
            max_row: Maximum row number
        """
        from openpyxl.utils import get_column_letter

        col_letter = get_column_letter(col_idx)
        range_string = f"{col_letter}2:{col_letter}{max_row}"

        # Green (low risk) to Red (high risk)
        rule = ColorScaleRule(
            start_type='num',
            start_value=0,
            start_color='63BE7B',  # Green
            mid_type='num',
            mid_value=50,
            mid_color='FFEB84',  # Yellow
            end_type='num',
            end_value=100,
            end_color='F8696B'  # Red
        )

        ws.conditional_formatting.add(range_string, rule)
