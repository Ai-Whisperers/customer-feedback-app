"""Styled Excel export strategy with formatting and charts."""

import io
from typing import Dict, Any, Tuple
from openpyxl import Workbook
from datetime import datetime

from app.services.export.base_exporter import BaseExporter
from app.services.export.sheet_builders.summary_builder import SummarySheetBuilder
from app.services.export.sheet_builders.details_builder import DetailsSheetBuilder
from app.services.export.sheet_builders.emotions_builder import EmotionsSheetBuilder
from app.services.export.sheet_builders.pain_points_builder import PainPointsSheetBuilder
from app.schemas.export import ExportInclude


class ExcelStyledExporter(BaseExporter):
    """
    Styled Excel export strategy.

    Creates multi-sheet Excel workbook with:
    - Professional styling
    - Conditional formatting
    - Charts and visualizations
    - Comprehensive metadata
    """

    def __init__(self):
        """Initialize styled Excel exporter with sheet builders."""
        self.summary_builder = SummarySheetBuilder()
        self.details_builder = DetailsSheetBuilder()
        self.emotions_builder = EmotionsSheetBuilder()
        self.pain_points_builder = PainPointsSheetBuilder()

    def export(
        self,
        results: Dict[str, Any],
        include: ExportInclude,
        task_id: str
    ) -> Tuple[bytes, str]:
        """
        Export results to styled Excel format.

        Args:
            results: Analysis results dictionary
            include: What to include (ALL, SUMMARY, DETAILED)
            task_id: Task identifier

        Returns:
            Tuple of (Excel content as bytes, filename)
        """
        # Validate results structure
        self._validate_results(results)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        rows_data = results.get("rows", [])

        # Build sheets based on include type
        if include == ExportInclude.SUMMARY:
            self.summary_builder.build(wb, results)
        elif include == ExportInclude.DETAILED and rows_data:
            self.details_builder.build(wb, rows_data)
        else:  # ExportInclude.ALL
            # Add all sheets
            self.summary_builder.build(wb, results)

            if rows_data:
                self.details_builder.build(wb, rows_data)
                self.emotions_builder.build(wb, rows_data)
                self.pain_points_builder.build(wb, results)

            # Add metadata sheet
            self._add_metadata_sheet(wb, results)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()

        # Generate filename
        filename = self._generate_filename(task_id, 'xlsx')

        return excel_content, filename

    def get_media_type(self) -> str:
        """
        Get MIME type for Excel files.

        Returns:
            Excel MIME type string
        """
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _add_metadata_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Add metadata sheet with export information.

        Args:
            wb: Workbook object
            results: Analysis results dictionary
        """
        from app.services.export.formatters.excel_styles import (
            apply_header_style, apply_body_style, ExcelColors
        )

        ws = wb.create_sheet("Metadatos")
        metadata = results.get("metadata", {})

        # Add title
        title_cell = ws.cell(row=1, column=1, value="Información de Exportación")
        ws.merge_cells('A1:B1')
        apply_header_style(title_cell, ExcelColors.DARK_GRAY)

        # Add metadata fields
        metadata_fields = [
            ("Total Comentarios", metadata.get("total_comments", 0)),
            ("Tiempo de Procesamiento (s)", round(metadata.get("processing_time_seconds", 0), 2)),
            ("Modelo Utilizado", metadata.get("model_used", "N/A")),
            ("Lotes Procesados", metadata.get("batches_processed", 0)),
            ("Fecha de Análisis", metadata.get("timestamp", "N/A")),
            ("Fecha de Exportación", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]

        # Add language distribution if available
        languages = metadata.get("languages", {})
        if languages:
            metadata_fields.append(("---", "---"))  # Separator
            metadata_fields.append(("Distribución de Idiomas", ""))
            for lang, count in languages.items():
                metadata_fields.append((f"  {lang.upper()}", count))

        for row_idx, (field, value) in enumerate(metadata_fields, 3):
            field_cell = ws.cell(row=row_idx, column=1, value=field)
            value_cell = ws.cell(row=row_idx, column=2, value=value)

            if field.startswith("  "):  # Indented sub-item
                apply_body_style(field_cell, bold=False)
            elif field == "---":  # Separator
                continue
            else:
                apply_body_style(field_cell, bold=True)

            apply_body_style(value_cell)

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
