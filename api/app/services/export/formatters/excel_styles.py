"""Reusable Excel styling constants and utilities."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelColors:
    """Color palette for Excel exports."""
    PRIMARY = "366092"  # Blue
    SECONDARY = "4472A8"  # Light blue
    SUCCESS = "70AD47"  # Green
    WARNING = "C65911"  # Orange
    DANGER = "FF6B6B"  # Red
    WHITE = "FFFFFF"
    LIGHT_GRAY = "F2F2F2"
    DARK_GRAY = "808080"


class ExcelFonts:
    """Reusable font definitions."""

    @staticmethod
    def header(color: str = ExcelColors.WHITE, size: int = 12) -> Font:
        """
        Standard header font.

        Args:
            color: Font color hex code
            size: Font size in points

        Returns:
            Font object for headers
        """
        return Font(bold=True, color=color, size=size)

    @staticmethod
    def body(bold: bool = False, size: int = 11) -> Font:
        """
        Standard body font.

        Args:
            bold: Whether font should be bold
            size: Font size in points

        Returns:
            Font object for body text
        """
        return Font(bold=bold, size=size)

    @staticmethod
    def metric(bold: bool = True, size: int = 14) -> Font:
        """
        Font for key metrics/numbers.

        Args:
            bold: Whether font should be bold
            size: Font size in points

        Returns:
            Font object for metrics
        """
        return Font(bold=bold, size=size)


class ExcelFills:
    """Reusable fill patterns."""

    @staticmethod
    def solid(color: str) -> PatternFill:
        """
        Solid color fill.

        Args:
            color: Hex color code

        Returns:
            PatternFill object
        """
        return PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

    @staticmethod
    def header_primary() -> PatternFill:
        """Primary header fill (dark blue)."""
        return ExcelFills.solid(ExcelColors.PRIMARY)

    @staticmethod
    def header_secondary() -> PatternFill:
        """Secondary header fill (light blue)."""
        return ExcelFills.solid(ExcelColors.SECONDARY)

    @staticmethod
    def header_success() -> PatternFill:
        """Success header fill (green)."""
        return ExcelFills.solid(ExcelColors.SUCCESS)

    @staticmethod
    def header_warning() -> PatternFill:
        """Warning header fill (orange)."""
        return ExcelFills.solid(ExcelColors.WARNING)


class ExcelAlignments:
    """Reusable alignment definitions."""

    @staticmethod
    def center() -> Alignment:
        """Center-aligned text."""
        return Alignment(horizontal="center", vertical="center")

    @staticmethod
    def left() -> Alignment:
        """Left-aligned text."""
        return Alignment(horizontal="left", vertical="center")

    @staticmethod
    def right() -> Alignment:
        """Right-aligned text."""
        return Alignment(horizontal="right", vertical="center")

    @staticmethod
    def center_wrapped() -> Alignment:
        """Center-aligned text with word wrap."""
        return Alignment(horizontal="center", vertical="center", wrap_text=True)


class ExcelBorders:
    """Reusable border definitions."""

    @staticmethod
    def thin(color: str = ExcelColors.DARK_GRAY) -> Border:
        """
        Thin border on all sides.

        Args:
            color: Border color hex code

        Returns:
            Border object
        """
        side = Side(style="thin", color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def header() -> Border:
        """Border for header cells."""
        side = Side(style="medium", color=ExcelColors.WHITE)
        return Border(bottom=side)


def apply_header_style(cell, color: str = ExcelColors.PRIMARY) -> None:
    """
    Apply consistent header styling to a cell.

    Args:
        cell: Openpyxl cell object
        color: Background color for header
    """
    cell.font = ExcelFonts.header()
    cell.fill = ExcelFills.solid(color)
    cell.alignment = ExcelAlignments.center()
    cell.border = ExcelBorders.thin()


def apply_body_style(cell, bold: bool = False) -> None:
    """
    Apply consistent body styling to a cell.

    Args:
        cell: Openpyxl cell object
        bold: Whether text should be bold
    """
    cell.font = ExcelFonts.body(bold=bold)
    cell.alignment = ExcelAlignments.left()
    cell.border = ExcelBorders.thin()


def auto_adjust_column_width(worksheet, max_width: int = 50) -> None:
    """
    Auto-adjust column widths based on content.

    Args:
        worksheet: Openpyxl worksheet object
        max_width: Maximum column width in characters
    """
    from openpyxl.cell.cell import MergedCell

    for column in worksheet.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            # Skip merged cells as they don't have column_letter
            if isinstance(cell, MergedCell):
                continue

            # Get column letter from first non-merged cell
            if column_letter is None:
                column_letter = cell.column_letter

            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Only adjust if we found a valid column letter
        if column_letter:
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
