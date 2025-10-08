"""Basic Excel export strategy without styling."""

import io
from typing import Dict, Any, Tuple
from openpyxl import Workbook
from datetime import datetime

from app.services.export.base_exporter import BaseExporter
from app.services.export.formatters.dataframe_formatter import DataFrameFormatter
from app.schemas.export import ExportInclude


class ExcelBasicExporter(BaseExporter):
    """
    Basic Excel export strategy.

    Creates multi-sheet Excel workbook without advanced formatting.
    Fast export suitable for large datasets.
    """

    def __init__(self):
        """Initialize basic Excel exporter."""
        self.formatter = DataFrameFormatter()

    def export(
        self,
        results: Dict[str, Any],
        include: ExportInclude,
        task_id: str
    ) -> Tuple[bytes, str]:
        """
        Export results to basic Excel format.

        Args:
            results: Analysis results dictionary
            include: What to include (ALL, SUMMARY, DETAILED)
            task_id: Task identifier

        Returns:
            Tuple of (Excel content as bytes, filename)
        """
        # Validate results structure
        self._validate_results(results)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        rows_data = results.get("rows", [])

        # Add sheets based on include type
        if include == ExportInclude.SUMMARY:
            self._add_summary_sheet(wb, results)
        elif include == ExportInclude.DETAILED and rows_data:
            self._add_detailed_sheet(wb, rows_data)
        else:  # ExportInclude.ALL
            self._add_summary_sheet(wb, results)
            if rows_data:
                self._add_detailed_sheet(wb, rows_data)
                self._add_emotions_sheet(wb, rows_data)
                self._add_pain_points_sheet(wb, results)
            self._add_metadata_sheet(wb, results)

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_content = excel_buffer.getvalue()

        # Generate filename
        filename = self._generate_filename(task_id, 'xlsx')

        return excel_content, filename

    def get_media_type(self) -> str:
        """
        Get MIME type for Excel files.

        Returns:
            Excel MIME type string
        """
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _add_summary_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Add summary sheet to workbook.

        Args:
            wb: Workbook object
            results: Analysis results
        """
        ws = wb.create_sheet("Resumen")
        df = self.formatter.create_summary_dataframe(results)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_detailed_sheet(self, wb: Workbook, rows_data: list) -> None:
        """
        Add detailed analysis sheet to workbook.

        Args:
            wb: Workbook object
            rows_data: List of row dictionaries
        """
        ws = wb.create_sheet("Análisis Detallado")
        df = self.formatter.create_detailed_dataframe(rows_data)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_emotions_sheet(self, wb: Workbook, rows_data: list) -> None:
        """
        Add emotions analysis sheet to workbook.

        Args:
            wb: Workbook object
            rows_data: List of row dictionaries
        """
        ws = wb.create_sheet("Emociones")
        df = self.formatter.create_emotions_dataframe(rows_data)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_pain_points_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Add pain points sheet to workbook.

        Args:
            wb: Workbook object
            results: Analysis results
        """
        ws = wb.create_sheet("Puntos de Dolor")
        df = self.formatter.create_pain_points_dataframe(results)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _add_metadata_sheet(self, wb: Workbook, results: Dict[str, Any]) -> None:
        """
        Add metadata sheet to workbook.

        Args:
            wb: Workbook object
            results: Analysis results
        """
        ws = wb.create_sheet("Metadatos")
        metadata = results.get("metadata", {})

        # Add metadata fields
        metadata_fields = [
            ("Total Comentarios", metadata.get("total_comments", 0)),
            ("Tiempo de Procesamiento (s)", round(metadata.get("processing_time_seconds", 0), 2)),
            ("Modelo Utilizado", metadata.get("model_used", "N/A")),
            ("Lotes Procesados", metadata.get("batches_processed", 0)),
            ("Fecha de Exportación", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]

        for row_idx, (field, value) in enumerate(metadata_fields, 1):
            ws.cell(row=row_idx, column=1, value=field)
            ws.cell(row=row_idx, column=2, value=value)
