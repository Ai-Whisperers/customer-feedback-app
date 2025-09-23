"""
Export service for generating CSV and Excel files.
Handles all export business logic.
"""

import io
from datetime import datetime
from typing import Dict, Any, Tuple
import pandas as pd
import structlog
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from app.config import settings
from app.schemas.export import ExportFormat, ExportInclude

logger = structlog.get_logger()


def generate_export(
    results: Dict[str, Any],
    format: ExportFormat,
    include: ExportInclude,
    task_id: str
) -> Tuple[bytes, str, str]:
    """
    Generate export file from analysis results.

    Args:
        results: Analysis results dictionary
        format: Export format (CSV or XLSX)
        include: What to include in export
        task_id: Task identifier

    Returns:
        Tuple of (file_content, filename, media_type)
    """
    if format == ExportFormat.CSV:
        file_content, filename = _generate_csv_export(results, include, task_id)
        media_type = "text/csv"
    else:  # XLSX
        file_content, filename = _generate_xlsx_export(results, include, task_id)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    logger.info(
        "Export generated",
        task_id=task_id,
        format=format.value,
        include=include.value,
        file_size_kb=len(file_content) / 1024
    )

    return file_content, filename, media_type


def _generate_csv_export(
    results: Dict[str, Any],
    include: ExportInclude,
    task_id: str
) -> Tuple[bytes, str]:
    """Generate CSV export from results."""
    df = _create_dataframe(results, include)

    # Convert to CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False, encoding='utf-8')
    csv_content = csv_buffer.getvalue().encode('utf-8')

    filename = f"analysis_{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return csv_content, filename


def _generate_xlsx_export(
    results: Dict[str, Any],
    include: ExportInclude,
    task_id: str
) -> Tuple[bytes, str]:
    """Generate Excel export with multiple sheets."""

    # Check if formatting is enabled
    if settings.EXCEL_FORMATTING_ENABLED:
        try:
            # Use the enhanced formatter
            from app.core.excel_formatter import create_formatter

            formatter = create_formatter()

            # Prepare data for formatter
            df = _create_detailed_dataframe(results.get("rows", []))
            aggregated_results = results.get("summary", {})
            metadata = results.get("metadata", {})
            metadata['processing_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Generate formatted workbook
            excel_buffer = formatter.format_analysis_workbook(
                df=df,
                aggregated_results=aggregated_results,
                metadata=metadata
            )

            logger.info("Excel export using enhanced formatter", task_id=task_id)

        except Exception as e:
            # Fallback to basic export
            logger.warning(
                "Failed to use enhanced formatter, falling back to basic",
                task_id=task_id,
                error=str(e)
            )
            excel_buffer = _generate_basic_xlsx(results, include)
    else:
        # Use basic export
        excel_buffer = _generate_basic_xlsx(results, include)

    excel_buffer.seek(0)
    excel_content = excel_buffer.read()

    filename = f"analysis_{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return excel_content, filename


def _generate_basic_xlsx(results: Dict[str, Any], include: ExportInclude) -> io.BytesIO:
    """Generate styled Excel export with charts."""
    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # Summary sheet with charts
        if include in [ExportInclude.ALL, ExportInclude.SUMMARY]:
            _add_styled_summary_sheet(writer, results)

        # Detailed analysis sheet with formatting
        if include in [ExportInclude.ALL, ExportInclude.DETAILED]:
            _add_styled_detailed_sheet(writer, results)

        # Additional sheets for ALL export
        if include == ExportInclude.ALL:
            _add_styled_emotions_sheet(writer, results)
            _add_styled_pain_points_sheet(writer, results)
            _add_metadata_sheet(writer, results)

    return excel_buffer


def _create_dataframe(results: Dict[str, Any], include: ExportInclude) -> pd.DataFrame:
    """Create dataframe based on include type."""
    rows_data = results.get("rows", [])

    if include == ExportInclude.SUMMARY or not rows_data:
        return _create_summary_dataframe(results)
    else:
        return _create_detailed_dataframe(rows_data)


def _create_summary_dataframe(results: Dict[str, Any]) -> pd.DataFrame:
    """Create summary dataframe."""
    summary = results.get("summary", {})
    nps_data = summary.get("nps", {})
    churn_data = summary.get("churn_risk", {})
    metadata = results.get("metadata", {})

    summary_data = {
        "Métrica": ["NPS Score", "Total Comentarios", "Promotores", "Pasivos", "Detractores",
                   "Riesgo Promedio (%)", "Alto Riesgo", "Tiempo Procesamiento (s)"],
        "Valor": [
            round(nps_data.get("score", 0), 1),
            metadata.get("total_comments", 0),
            f"{nps_data.get('promoters', 0)} ({round(nps_data.get('promoters_percentage', 0), 1)}%)",
            f"{nps_data.get('passives', 0)} ({round(nps_data.get('passives_percentage', 0), 1)}%)",
            f"{nps_data.get('detractors', 0)} ({round(nps_data.get('detractors_percentage', 0), 1)}%)",
            round(churn_data.get("average", 0) * 100, 1),
            f"{churn_data.get('high_risk_count', 0)} ({round(churn_data.get('high_risk_percentage', 0), 1)}%)",
            round(metadata.get("processing_time_seconds", 0), 1)
        ]
    }
    return pd.DataFrame(summary_data)


def _create_detailed_dataframe(rows_data: list) -> pd.DataFrame:
    """Create detailed dataframe from row data."""
    detailed_data = []

    for row in rows_data:
        row_data = {
            "Index": row.get("index", 0) + 1,  # 1-based index for Excel
            "Comentario": row.get("original_text", ""),
            "Nota": row.get("nota", 5),
            "Categoría NPS": row.get("nps_category", "passive"),
            "Sentimiento": row.get("sentiment", "neutral"),
            "Idioma": row.get("language", "es"),
            "Riesgo de Abandono (%)": round(row.get("churn_risk", 0.5) * 100, 1),
            "Puntos de Dolor": "; ".join(row.get("pain_points", [])) if row.get("pain_points") else "N/A"
        }

        # Add top emotions only (above 0.3 threshold)
        emotions = row.get("emotions", {})
        top_emotions = {k: v for k, v in emotions.items() if v > 0.3}
        if top_emotions:
            sorted_emotions = sorted(top_emotions.items(), key=lambda x: x[1], reverse=True)[:3]
            row_data["Emociones Principales"] = ", ".join([f"{e[0].replace('_', ' ').title()}: {round(e[1]*100)}%" for e in sorted_emotions])
        else:
            row_data["Emociones Principales"] = "N/A"

        detailed_data.append(row_data)

    return pd.DataFrame(detailed_data)


def _add_summary_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add summary sheet to Excel file."""
    summary = results.get("summary", {})
    nps_data = summary.get("nps", {})
    churn_data = summary.get("churn_risk", {})

    summary_df = pd.DataFrame([
        {"Metric": "NPS Score", "Value": nps_data.get("score")},
        {"Metric": "Total Comments", "Value": results.get("metadata", {}).get("total_comments")},
        {"Metric": "Promoters", "Value": f"{nps_data.get('promoters')} ({nps_data.get('promoters_percentage')}%)"},
        {"Metric": "Passives", "Value": f"{nps_data.get('passives')} ({nps_data.get('passives_percentage')}%)"},
        {"Metric": "Detractors", "Value": f"{nps_data.get('detractors')} ({nps_data.get('detractors_percentage')}%)"},
        {"Metric": "Average Churn Risk", "Value": f"{churn_data.get('average', 0):.2%}"},
        {"Metric": "High Risk Customers", "Value": f"{churn_data.get('high_risk_count')} ({churn_data.get('high_risk_percentage')}%)"}
    ])

    summary_df.to_excel(writer, sheet_name='Summary', index=False)


def _add_detailed_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add detailed analysis sheet to Excel file."""
    rows_data = results.get("rows", [])

    if rows_data:
        detailed_data = []
        for row in rows_data:
            row_data = {
                "Index": row.get("index", 0) + 1,
                "Comentario": row.get("original_text", "")[:500],  # More text for Excel
                "Nota": row.get("nota", 5),
                "NPS": row.get("nps_category", "passive"),
                "Sentimiento": row.get("sentiment", "neutral"),
                "Idioma": row.get("language", "es"),
                "Riesgo (%)": round(row.get("churn_risk", 0.5) * 100, 1),
                "Puntos de Dolor": "; ".join(row.get("pain_points", [])) if row.get("pain_points") else ""
            }

            # Add top emotions
            emotions = row.get("emotions", {})
            top_emotions = sorted([(k, v) for k, v in emotions.items() if v > 0.2], key=lambda x: x[1], reverse=True)[:3]
            if top_emotions:
                row_data["Emociones"] = ", ".join([f"{e[0].replace('_', ' ')}: {round(e[1]*100)}%" for e in top_emotions])
            else:
                row_data["Emociones"] = ""

            detailed_data.append(row_data)

        detailed_df = pd.DataFrame(detailed_data)
        detailed_df.to_excel(writer, sheet_name='Detailed', index=False)


def _add_emotions_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add emotions matrix sheet to Excel file."""
    rows_data = results.get("rows", [])

    if rows_data:
        emotions_data = []
        for row in rows_data:
            emotion_row = {"Index": row.get("index")}
            emotions = row.get("emotions", {})
            emotion_row.update(emotions)
            emotions_data.append(emotion_row)

        emotions_df = pd.DataFrame(emotions_data)
        emotions_df.to_excel(writer, sheet_name='Emotions', index=False)


def _add_pain_points_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add pain points sheet to Excel file."""
    pain_points = results.get("summary", {}).get("pain_points", [])

    if pain_points:
        pain_df = pd.DataFrame(pain_points)
        pain_df.to_excel(writer, sheet_name='Pain Points', index=False)


def _add_metadata_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add metadata sheet to Excel file."""
    metadata = results.get("metadata", {})
    meta_df = pd.DataFrame([{
        "Processing Time": f"{metadata.get('processing_time_seconds', 0):.2f} seconds",
        "Model Used": metadata.get("model_used"),
        "Total Comments": metadata.get("total_comments"),
        "Timestamp": metadata.get("timestamp"),
        "Batches Processed": metadata.get("batches_processed")
    }])
    meta_df.to_excel(writer, sheet_name='Metadata', index=False)


def _add_styled_summary_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add styled summary sheet with charts to Excel file."""
    summary_df = _create_summary_dataframe(results)
    summary_df.to_excel(writer, sheet_name='Resumen', index=False)

    # Get worksheet for styling
    worksheet = writer.sheets['Resumen']

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Style headers
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Adjust column widths
    for col in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 25

    # Add NPS distribution pie chart
    summary = results.get("summary", {})
    nps_data = summary.get("nps", {})

    if nps_data:
        # Create data for chart
        chart_data = [
            ["Categoría", "Cantidad"],
            ["Promotores", nps_data.get("promoters", 0)],
            ["Pasivos", nps_data.get("passives", 0)],
            ["Detractores", nps_data.get("detractors", 0)]
        ]

        # Write chart data
        start_row = worksheet.max_row + 3
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                worksheet.cell(row=start_row + row_idx, column=col_idx + 1, value=value)

        # Create pie chart
        pie = PieChart()
        pie.title = "Distribución NPS"
        pie.add_data(Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + 3), titles_from_data=False)
        pie.set_categories(Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + 3))
        pie.width = 10
        pie.height = 8

        worksheet.add_chart(pie, f"D{start_row}")


def _add_styled_detailed_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add styled detailed sheet to Excel file."""
    rows_data = results.get("rows", [])

    if rows_data:
        detailed_data = []
        for row in rows_data:
            row_data = {
                "Index": row.get("index", 0) + 1,
                "Comentario": row.get("original_text", "")[:500],
                "Nota": row.get("nota", 5),
                "NPS": row.get("nps_category", "passive"),
                "Riesgo (%)": round(row.get("churn_risk", 0.5) * 100, 1),
                "Puntos de Dolor": "; ".join(row.get("pain_points", [])) if row.get("pain_points") else ""
            }

            # Add top emotions
            emotions = row.get("emotions", {})
            top_emotions = sorted([(k, v) for k, v in emotions.items() if v > 0.2], key=lambda x: x[1], reverse=True)[:3]
            if top_emotions:
                row_data["Emociones"] = ", ".join([f"{e[0].replace('_', ' ')}: {round(e[1]*100)}%" for e in top_emotions])
            else:
                row_data["Emociones"] = ""

            detailed_data.append(row_data)

        detailed_df = pd.DataFrame(detailed_data)
        detailed_df.to_excel(writer, sheet_name='Detalle', index=False)

        # Style the sheet
        worksheet = writer.sheets['Detalle']

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472A8", end_color="4472A8", fill_type="solid")

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply conditional formatting for risk column
        for row in range(2, worksheet.max_row + 1):
            risk_cell = worksheet.cell(row=row, column=5)  # Risk column
            risk_value = risk_cell.value
            if risk_value and isinstance(risk_value, (int, float)):
                if risk_value >= 60:
                    risk_cell.fill = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
                elif risk_value >= 40:
                    risk_cell.fill = PatternFill(start_color="FFFFB3", end_color="FFFFB3", fill_type="solid")
                else:
                    risk_cell.fill = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 8
        worksheet.column_dimensions['B'].width = 60
        worksheet.column_dimensions['C'].width = 8
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 12
        worksheet.column_dimensions['F'].width = 30
        worksheet.column_dimensions['G'].width = 40


def _add_styled_emotions_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add styled emotions sheet with chart to Excel file."""
    rows_data = results.get("rows", [])

    if rows_data:
        # Calculate emotion averages
        emotion_totals = {}
        for row in rows_data:
            for emotion, value in row.get("emotions", {}).items():
                if emotion not in emotion_totals:
                    emotion_totals[emotion] = []
                emotion_totals[emotion].append(value)

        emotion_averages = {
            emotion: sum(values) / len(values) * 100
            for emotion, values in emotion_totals.items()
        }

        # Sort by average value
        sorted_emotions = sorted(emotion_averages.items(), key=lambda x: x[1], reverse=True)[:10]

        emotion_df = pd.DataFrame(sorted_emotions, columns=["Emoción", "Promedio (%)"])
        emotion_df.to_excel(writer, sheet_name='Emociones', index=False)

        # Style worksheet
        worksheet = writer.sheets['Emociones']

        # Style headers
        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 15

        # Add bar chart
        chart = BarChart()
        chart.title = "Top 10 Emociones"
        chart.x_axis.title = "Emoción"
        chart.y_axis.title = "Promedio (%)"
        chart.width = 12
        chart.height = 8

        data = Reference(worksheet, min_col=2, min_row=1, max_row=11)
        categories = Reference(worksheet, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        worksheet.add_chart(chart, "D2")


def _add_styled_pain_points_sheet(writer: pd.ExcelWriter, results: Dict[str, Any]):
    """Add styled pain points sheet to Excel file."""
    pain_points = results.get("summary", {}).get("pain_points", [])

    # If no pain_points in summary, aggregate from rows data (like UI does)
    if not pain_points and results.get("rows"):
        pain_points_counter = {}
        total_comments = len(results["rows"])

        for row in results["rows"]:
            for pain_point in row.get("pain_points", []):
                pain_points_counter[pain_point] = pain_points_counter.get(pain_point, 0) + 1

        # Convert to list format
        pain_points = [
            {
                "issue": issue,
                "frequency": freq,
                "percentage": round(freq / total_comments * 100, 1) if total_comments > 0 else 0
            }
            for issue, freq in sorted(pain_points_counter.items(), key=lambda x: x[1], reverse=True)[:10]
        ]

    if pain_points:
        # Transform pain_points to proper format for Excel
        formatted_pain_points = []
        for point in pain_points:
            formatted_pain_points.append({
                "Punto de Dolor": point.get("issue", point.get("category", "N/A")),
                "Frecuencia": point.get("frequency", point.get("count", 0)),
                "Porcentaje": f"{point.get('percentage', 0)}%",
                "Ejemplos": ", ".join(point.get("examples", [])) if point.get("examples") else "N/A"
            })

        if formatted_pain_points:
            pain_df = pd.DataFrame(formatted_pain_points)
        else:
            # Create empty dataframe with proper columns if no data
            pain_df = pd.DataFrame(columns=["Punto de Dolor", "Frecuencia", "Porcentaje", "Ejemplos"])

        pain_df.to_excel(writer, sheet_name='Puntos de Dolor', index=False)
    else:
        # Create empty dataframe with message
        pain_df = pd.DataFrame([{
            "Punto de Dolor": "No se encontraron puntos de dolor",
            "Frecuencia": 0,
            "Porcentaje": "0%",
            "Ejemplos": "N/A"
        }])
        pain_df.to_excel(writer, sheet_name='Puntos de Dolor', index=False)

    # Style worksheet
    worksheet = writer.sheets['Puntos de Dolor']

    # Style headers
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width