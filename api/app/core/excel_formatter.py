"""
Excel formatter module for enhanced export styling.
Provides professional formatting for customer feedback analysis results.
"""

import io
from typing import Dict, Any, Optional, List
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import structlog

from app.config import settings

logger = structlog.get_logger()


class ExcelFormatter:
    """Professional Excel formatter for analysis results."""

    # Color scheme
    COLORS = {
        'header': 'FF2E3440',  # Dark blue-gray
        'subheader': 'FF4A5568',  # Medium gray
        'positive': 'FF48BB78',  # Green
        'negative': 'FFF56565',  # Red
        'neutral': 'FFED8936',  # Orange
        'row_even': 'FFF7FAFC',  # Light gray
        'row_odd': 'FFFFFFFF',  # White
        'border': 'FFE2E8F0'  # Border gray
    }

    # Column width mappings
    COLUMN_WIDTHS = {
        'index': 8,
        'rating': 10,
        'nps': 15,
        'comment': 50,
        'emotions': 12,
        'churn_risk': 12,
        'pain_points': 30,
        'default': 15
    }

    def __init__(self, enable_charts: bool = True, enable_conditional: bool = True):
        """
        Initialize formatter.

        Args:
            enable_charts: Enable chart generation
            enable_conditional: Enable conditional formatting
        """
        self.enable_charts = enable_charts
        self.enable_conditional = enable_conditional
        self._create_styles()

    def _create_styles(self):
        """Create reusable named styles."""
        # Header style
        self.header_style = NamedStyle(name='header_style')
        self.header_style.font = Font(bold=True, color='FFFFFFFF', size=12)
        self.header_style.fill = PatternFill(
            start_color=self.COLORS['header'],
            end_color=self.COLORS['header'],
            fill_type='solid'
        )
        self.header_style.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        self.header_style.border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='medium', color=self.COLORS['border']),
            bottom=Side(style='medium', color=self.COLORS['border'])
        )

    def format_analysis_workbook(
        self,
        df: pd.DataFrame,
        aggregated_results: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> io.BytesIO:
        """
        Create formatted Excel workbook with multiple sheets.

        Args:
            df: Main data DataFrame
            aggregated_results: Aggregated analysis results
            metadata: Processing metadata

        Returns:
            BytesIO object with Excel file
        """
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(wb, aggregated_results, metadata)
        self._create_details_sheet(wb, df)
        self._create_emotions_sheet(wb, aggregated_results)
        self._create_pain_points_sheet(wb, aggregated_results)

        if self.enable_charts:
            self._create_charts_sheet(wb, aggregated_results)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_summary_sheet(
        self,
        wb: Workbook,
        aggregated_results: Dict[str, Any],
        metadata: Dict[str, Any]
    ):
        """Create summary dashboard sheet."""
        ws = wb.create_sheet("Resumen Ejecutivo")

        # Title
        ws['A1'] = "Análisis de Feedback - Resumen Ejecutivo"
        ws['A1'].font = Font(bold=True, size=16, color=self.COLORS['header'])
        ws.merge_cells('A1:F1')

        # Metadata section
        row = 3
        ws[f'A{row}'] = "Información del Análisis"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metadata_items = [
            ("Total de comentarios", metadata.get('total_comments', 0)),
            ("Idioma predominante", self._get_language_label(metadata.get('language', 'es'))),
            ("Fecha de procesamiento", metadata.get('processing_date', 'N/A')),
            ("Tiempo de procesamiento", f"{metadata.get('processing_time_seconds', 0):.1f}s")
        ]

        for label, value in metadata_items:
            ws[f'B{row}'] = label
            ws[f'C{row}'] = value
            row += 1

        # NPS section
        row += 1
        ws[f'A{row}'] = "Net Promoter Score (NPS)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        nps_data = aggregated_results.get('nps', {})
        nps_score = nps_data.get('score', 0)

        ws[f'B{row}'] = "Score NPS"
        ws[f'C{row}'] = nps_score
        self._apply_nps_color(ws[f'C{row}'], nps_score)
        row += 1

        ws[f'B{row}'] = "Promotores"
        ws[f'C{row}'] = f"{nps_data.get('promoters_percentage', 0):.1f}%"
        ws[f'C{row}'].font = Font(color=self.COLORS['positive'])
        row += 1

        ws[f'B{row}'] = "Pasivos"
        ws[f'C{row}'] = f"{nps_data.get('passives_percentage', 0):.1f}%"
        ws[f'C{row}'].font = Font(color=self.COLORS['neutral'])
        row += 1

        ws[f'B{row}'] = "Detractores"
        ws[f'C{row}'] = f"{nps_data.get('detractors_percentage', 0):.1f}%"
        ws[f'C{row}'].font = Font(color=self.COLORS['negative'])
        row += 1

        # Churn Risk section
        row += 1
        churn_data = aggregated_results.get('churn_risk', {})
        ws[f'A{row}'] = "Riesgo de Abandono"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'B{row}'] = "Riesgo promedio"
        ws[f'C{row}'] = f"{churn_data.get('average', 0):.2f}"
        row += 1

        ws[f'B{row}'] = "Clientes en alto riesgo"
        ws[f'C{row}'] = f"{churn_data.get('high_risk_percentage', 0):.1f}%"
        ws[f'C{row}'].font = Font(color=self.COLORS['negative'])

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15

    def _create_details_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create detailed results sheet."""
        ws = wb.create_sheet("Detalle de Comentarios")

        # Write headers
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.style = self.header_style

        # Write data with formatting
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Apply alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=self.COLORS['row_even'],
                        end_color=self.COLORS['row_even'],
                        fill_type='solid'
                    )

                # Format specific columns
                col_name = headers[col_idx - 1].lower()
                if 'nps' in col_name:
                    self._format_nps_cell(cell, value)
                elif 'churn' in col_name or 'riesgo' in col_name:
                    self._format_risk_cell(cell, value)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws, df)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add conditional formatting for numeric columns
        if self.enable_conditional:
            self._add_conditional_formatting(ws, df)

    def _create_emotions_sheet(
        self,
        wb: Workbook,
        aggregated_results: Dict[str, Any]
    ):
        """Create emotions analysis sheet."""
        ws = wb.create_sheet("Análisis de Emociones")

        emotions = aggregated_results.get('emotions', {})
        if not emotions:
            ws['A1'] = "No hay datos de emociones disponibles"
            return

        # Title
        ws['A1'] = "Distribución de Emociones"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['A3'] = "Emoción"
        ws['B3'] = "Promedio"
        ws['C3'] = "Categoría"

        for cell in ['A3', 'B3', 'C3']:
            ws[cell].style = self.header_style

        # Emotion data
        row = 4
        sorted_emotions = sorted(emotions.items(), key=lambda x: x[1], reverse=True)

        for emotion, score in sorted_emotions:
            ws[f'A{row}'] = self._translate_emotion(emotion)
            ws[f'B{row}'] = f"{score:.3f}"

            # Categorize emotion
            if score > 0.5:
                category = "Alta"
                color = self.COLORS['positive']
            elif score > 0.2:
                category = "Media"
                color = self.COLORS['neutral']
            else:
                category = "Baja"
                color = self.COLORS['negative']

            ws[f'C{row}'] = category
            ws[f'C{row}'].font = Font(color=color)

            # Add data bar
            if self.enable_conditional:
                ws[f'B{row}'].fill = self._get_gradient_fill(score)

            row += 1

        # Adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_pain_points_sheet(
        self,
        wb: Workbook,
        aggregated_results: Dict[str, Any]
    ):
        """Create pain points analysis sheet."""
        ws = wb.create_sheet("Puntos de Dolor")

        pain_points = aggregated_results.get('pain_points', [])
        if not pain_points:
            ws['A1'] = "No se identificaron puntos de dolor"
            return

        # Title
        ws['A1'] = "Principales Puntos de Dolor"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['A3'] = "Punto de Dolor"
        ws['B3'] = "Frecuencia"
        ws['C3'] = "Porcentaje"

        for cell in ['A3', 'B3', 'C3']:
            ws[cell].style = self.header_style

        # Pain points data
        row = 4
        total_mentions = sum(pp.get('count', 0) for pp in pain_points[:10])

        for pp in pain_points[:10]:  # Top 10
            ws[f'A{row}'] = pp.get('pain_point', 'N/A')
            count = pp.get('count', 0)
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{(count/total_mentions*100):.1f}%" if total_mentions > 0 else "0%"
            row += 1

        # Adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_charts_sheet(
        self,
        wb: Workbook,
        aggregated_results: Dict[str, Any]
    ):
        """Create visual charts sheet."""
        ws = wb.create_sheet("Visualizaciones")

        # NPS Distribution Pie Chart
        self._add_nps_chart(ws, aggregated_results.get('nps', {}))

        # Emotions Bar Chart
        self._add_emotions_chart(ws, aggregated_results.get('emotions', {}))

    def _add_nps_chart(self, ws, nps_data: Dict[str, Any]):
        """Add NPS distribution pie chart."""
        # Prepare data
        ws['A1'] = "NPS Distribution"
        ws['A2'] = "Categoría"
        ws['B2'] = "Porcentaje"

        categories = ['Promotores', 'Pasivos', 'Detractores']
        percentages = [
            nps_data.get('promoters_percentage', 0),
            nps_data.get('passives_percentage', 0),
            nps_data.get('detractors_percentage', 0)
        ]

        for i, (cat, pct) in enumerate(zip(categories, percentages), 3):
            ws[f'A{i}'] = cat
            ws[f'B{i}'] = pct

        # Create pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=3, max_row=5)
        data = Reference(ws, min_col=2, min_row=2, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribución NPS"
        ws.add_chart(pie, "D2")

    def _add_emotions_chart(self, ws, emotions: Dict[str, float]):
        """Add emotions bar chart."""
        if not emotions:
            return

        # Prepare data
        ws['A10'] = "Top Emociones"
        ws['A11'] = "Emoción"
        ws['B11'] = "Valor"

        sorted_emotions = sorted(emotions.items(), key=lambda x: x[1], reverse=True)[:5]

        for i, (emotion, value) in enumerate(sorted_emotions, 12):
            ws[f'A{i}'] = self._translate_emotion(emotion)
            ws[f'B{i}'] = value

        # Create bar chart
        chart = BarChart()
        data = Reference(ws, min_col=2, min_row=11, max_row=16)
        categories = Reference(ws, min_col=1, min_row=12, max_row=16)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = "Top 5 Emociones"
        chart.x_axis.title = "Emoción"
        chart.y_axis.title = "Valor Promedio"
        ws.add_chart(chart, "D10")

    def _auto_adjust_columns(self, ws, df: pd.DataFrame):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width with limits
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 8)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_conditional_formatting(self, ws, df: pd.DataFrame):
        """Add conditional formatting rules."""
        # Find numeric columns
        for col_idx, col_name in enumerate(df.columns, 1):
            if df[col_name].dtype in ['float64', 'int64']:
                col_letter = get_column_letter(col_idx)

                # Skip if it's index or id
                if 'id' in col_name.lower() or 'index' in col_name.lower():
                    continue

                # Apply color scale for risk/churn columns
                if 'risk' in col_name.lower() or 'churn' in col_name.lower():
                    ws.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{len(df)+1}',
                        ColorScaleRule(
                            start_type='min', start_color='FF48BB78',  # Green
                            mid_type='percentile', mid_value=50, mid_color='FFED8936',  # Orange
                            end_type='max', end_color='FFF56565'  # Red
                        )
                    )

    def _format_nps_cell(self, cell, value):
        """Format NPS category cell."""
        if isinstance(value, str):
            value_lower = value.lower()
            if 'promoter' in value_lower or 'promotor' in value_lower:
                cell.font = Font(color=self.COLORS['positive'])
            elif 'passive' in value_lower or 'pasivo' in value_lower:
                cell.font = Font(color=self.COLORS['neutral'])
            elif 'detractor' in value_lower:
                cell.font = Font(color=self.COLORS['negative'])

    def _format_risk_cell(self, cell, value):
        """Format risk score cell."""
        try:
            risk_value = float(value)
            if risk_value >= 0.7:
                cell.font = Font(color=self.COLORS['negative'])
            elif risk_value >= 0.4:
                cell.font = Font(color=self.COLORS['neutral'])
            else:
                cell.font = Font(color=self.COLORS['positive'])
        except (ValueError, TypeError):
            pass

    def _apply_nps_color(self, cell, score):
        """Apply color based on NPS score."""
        if score >= 70:
            cell.font = Font(bold=True, color=self.COLORS['positive'])
        elif score >= 50:
            cell.font = Font(bold=True, color=self.COLORS['neutral'])
        else:
            cell.font = Font(bold=True, color=self.COLORS['negative'])

    def _get_gradient_fill(self, value: float) -> PatternFill:
        """Get gradient fill based on value."""
        # Convert value (0-1) to RGB gradient
        if value > 0.5:
            # Green gradient
            intensity = int(255 * (value - 0.5) * 2)
            color = f'FF{72:02X}{187 + intensity//4:02X}{120:02X}'
        else:
            # Red gradient
            intensity = int(255 * (1 - value * 2))
            color = f'FF{245 + intensity//25:02X}{101 - intensity//5:02X}{101:02X}'

        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _translate_emotion(self, emotion: str) -> str:
        """Translate emotion names to Spanish."""
        translations = {
            'joy': 'Alegría',
            'trust': 'Confianza',
            'fear': 'Miedo',
            'surprise': 'Sorpresa',
            'sadness': 'Tristeza',
            'disgust': 'Disgusto',
            'anger': 'Enojo',
            'anticipation': 'Anticipación',
            'love': 'Amor',
            'optimism': 'Optimismo',
            'pessimism': 'Pesimismo',
            'hope': 'Esperanza',
            'anxiety': 'Ansiedad',
            'envy': 'Envidia',
            'guilt': 'Culpa',
            'pride': 'Orgullo'
        }
        return translations.get(emotion, emotion.capitalize())

    def _get_language_label(self, lang: str) -> str:
        """Get language label."""
        languages = {
            'es': 'Español',
            'en': 'Inglés',
            'mixed': 'Mixto'
        }
        return languages.get(lang, lang.upper())


def create_formatter() -> ExcelFormatter:
    """Factory function to create formatter."""
    enable_charts = getattr(settings, 'EXCEL_ENABLE_CHARTS', True)
    enable_conditional = getattr(settings, 'EXCEL_ENABLE_CONDITIONAL', True)

    return ExcelFormatter(
        enable_charts=enable_charts,
        enable_conditional=enable_conditional
    )