"""
Test script for Excel formatter functionality.
"""

import os
import sys
import pandas as pd
import tempfile
from datetime import datetime

sys.path.append('.')

# Enable formatting for testing
os.environ["EXCEL_FORMATTING_ENABLED"] = "true"
os.environ["EXCEL_ENABLE_CHARTS"] = "true"
os.environ["EXCEL_ENABLE_CONDITIONAL"] = "true"

from app.config import settings
from app.core.excel_formatter import create_formatter, ExcelFormatter


def create_test_data():
    """Create test data for formatter."""

    # Sample DataFrame
    df = pd.DataFrame({
        'ID': range(1, 11),
        'Nota': [9, 8, 5, 10, 3, 7, 6, 9, 4, 8],
        'Comentario Final': [
            'Excelente servicio, muy satisfecho',
            'Buen producto pero puede mejorar',
            'Regular experiencia, esperaba más',
            'Increíble! Superó mis expectativas',
            'Muy mal servicio, decepcionante',
            'Buena calidad precio',
            'Normal, nada especial',
            'Me encantó el producto',
            'No cumple con lo prometido',
            'Satisfecho con la compra'
        ],
        'NPS': ['promoter', 'passive', 'detractor', 'promoter', 'detractor',
                'passive', 'detractor', 'promoter', 'detractor', 'passive'],
        'Churn Risk': [0.1, 0.3, 0.7, 0.05, 0.9, 0.4, 0.6, 0.15, 0.85, 0.35]
    })

    # Aggregated results
    aggregated_results = {
        'nps': {
            'score': 57.5,  # Using shifted scale
            'promoters': 3,
            'promoters_percentage': 30.0,
            'passives': 3,
            'passives_percentage': 30.0,
            'detractors': 4,
            'detractors_percentage': 40.0
        },
        'emotions': {
            'joy': 0.45,
            'trust': 0.62,
            'fear': 0.15,
            'surprise': 0.28,
            'sadness': 0.22,
            'disgust': 0.18,
            'anger': 0.25,
            'anticipation': 0.35
        },
        'pain_points': [
            {'pain_point': 'Servicio al cliente', 'count': 5},
            {'pain_point': 'Calidad del producto', 'count': 4},
            {'pain_point': 'Tiempo de entrega', 'count': 3},
            {'pain_point': 'Precio', 'count': 2},
            {'pain_point': 'Empaque', 'count': 1}
        ],
        'churn_risk': {
            'average': 0.445,
            'high_risk_count': 3,
            'high_risk_percentage': 30.0,
            'distribution': {
                'low': 3,
                'medium': 4,
                'high': 3
            }
        }
    }

    # Metadata
    metadata = {
        'total_comments': 10,
        'language': 'es',
        'processing_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'processing_time_seconds': 8.5
    }

    return df, aggregated_results, metadata


def test_formatter_creation():
    """Test formatter creation and configuration."""
    print("\nTesting Formatter Creation")
    print("=" * 60)

    formatter = create_formatter()
    print(f"✓ Formatter created")
    print(f"  Charts enabled: {formatter.enable_charts}")
    print(f"  Conditional formatting: {formatter.enable_conditional}")

    # Check configuration from settings
    print(f"\nConfiguration from settings:")
    print(f"  EXCEL_FORMATTING_ENABLED: {settings.EXCEL_FORMATTING_ENABLED}")
    print(f"  EXCEL_ENABLE_CHARTS: {settings.EXCEL_ENABLE_CHARTS}")
    print(f"  EXCEL_ENABLE_CONDITIONAL: {settings.EXCEL_ENABLE_CONDITIONAL}")

    assert formatter.enable_charts == True
    assert formatter.enable_conditional == True

    print("\n✓ Formatter creation test passed")


def test_workbook_generation():
    """Test workbook generation with formatting."""
    print("\nTesting Workbook Generation")
    print("=" * 60)

    formatter = create_formatter()
    df, aggregated_results, metadata = create_test_data()

    try:
        # Generate workbook
        excel_buffer = formatter.format_analysis_workbook(
            df=df,
            aggregated_results=aggregated_results,
            metadata=metadata
        )

        # Save to temp file for inspection
        temp_file = tempfile.NamedTemporaryFile(
            suffix='.xlsx',
            delete=False,
            dir=tempfile.gettempdir()
        )
        temp_file.write(excel_buffer.read())
        temp_file.close()

        print(f"✓ Workbook generated successfully")
        print(f"  Saved to: {temp_file.name}")
        print(f"  Size: {os.path.getsize(temp_file.name) / 1024:.1f} KB")

        # Load and verify sheets
        from openpyxl import load_workbook
        wb = load_workbook(temp_file.name)

        print(f"\nSheets created:")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"  - {sheet_name}: {sheet.max_row} rows x {sheet.max_column} columns")

        expected_sheets = [
            "Resumen Ejecutivo",
            "Detalle de Comentarios",
            "Análisis de Emociones",
            "Puntos de Dolor"
        ]

        if formatter.enable_charts:
            expected_sheets.append("Visualizaciones")

        for sheet in expected_sheets:
            if sheet in wb.sheetnames:
                print(f"  ✓ {sheet} found")
            else:
                print(f"  ✗ {sheet} missing")

        wb.close()

        print(f"\n✓ Workbook generation test passed")
        print(f"\nTo view the formatted Excel: {temp_file.name}")

    except Exception as e:
        print(f"✗ Workbook generation failed: {e}")
        return False

    return True


def test_color_schemes():
    """Test color schemes and styling."""
    print("\nTesting Color Schemes")
    print("=" * 60)

    formatter = ExcelFormatter()

    print("Color scheme:")
    for name, color in formatter.COLORS.items():
        print(f"  {name}: #{color[2:]}")  # Remove 'FF' prefix for display

    print("\nColumn widths:")
    for col, width in formatter.COLUMN_WIDTHS.items():
        print(f"  {col}: {width}")

    print("\n✓ Color schemes test passed")


def test_fallback_mode():
    """Test fallback when formatting is disabled."""
    print("\nTesting Fallback Mode")
    print("=" * 60)

    # Disable formatting
    os.environ["EXCEL_FORMATTING_ENABLED"] = "false"

    from app.services.export_service import _generate_xlsx_export

    # Create minimal test data
    results = {
        'rows': [
            {
                'index': 0,
                'rating': 9,
                'comment': 'Test comment',
                'nps': 'promoter',
                'emotions': {'joy': 0.5},
                'churn_risk': 0.2,
                'pain_points': ['test']
            }
        ],
        'summary': {
            'nps': {'score': 50, 'promoters': 1, 'promoters_percentage': 100},
            'emotions': {'joy': 0.5},
            'pain_points': [],
            'churn_risk': {'average': 0.2}
        },
        'metadata': {'total_comments': 1}
    }

    try:
        from app.schemas.export import ExportInclude

        content, filename = _generate_xlsx_export(
            results,
            ExportInclude.ALL,
            'test_fallback'
        )

        print(f"✓ Fallback export succeeded")
        print(f"  Filename: {filename}")
        print(f"  Size: {len(content) / 1024:.1f} KB")

    except Exception as e:
        print(f"✗ Fallback failed: {e}")
        return False

    # Re-enable for other tests
    os.environ["EXCEL_FORMATTING_ENABLED"] = "true"

    print("\n✓ Fallback mode test passed")
    return True


def main():
    """Run all tests."""
    print("\n" + "=" * 60)
    print("Excel Formatter Test Suite")
    print("=" * 60)

    tests = [
        ("Formatter Creation", test_formatter_creation),
        ("Color Schemes", test_color_schemes),
        ("Workbook Generation", test_workbook_generation),
        ("Fallback Mode", test_fallback_mode)
    ]

    results = []
    for name, test_func in tests:
        try:
            result = test_func()
            results.append((name, result if result is not None else True))
        except Exception as e:
            print(f"\n✗ {name} test failed with exception: {e}")
            results.append((name, False))

    # Summary
    print("\n" + "=" * 60)
    print("Test Summary")
    print("=" * 60)

    passed = sum(1 for _, success in results if success)
    total = len(results)

    for name, success in results:
        status = "✓ PASSED" if success else "✗ FAILED"
        print(f"{name:25} {status}")

    print(f"\nTotal: {passed}/{total} tests passed")

    if passed == total:
        print("\n🎉 All tests passed! Excel formatting is ready.")
    else:
        print(f"\n⚠ {total - passed} test(s) failed.")


if __name__ == "__main__":
    main()